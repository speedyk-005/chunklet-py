import csv
from pathlib import Path
from typing import Union, List, Optional, Callable, Dict, Any, Tuple
from box import Box
from loguru import logger
try:
    from striprtf.striprtf import rtf_to_text
except ImportError:
    rtf_to_text = None
from chunklet.plain_text_chunker import PlainTextChunker
from chunklet.utils.pdf_processor import PDFProcessor
from chunklet.utils.docx_processor import DOCXProcessor
from chunklet.utils.rst_to_md import rst_to_markdown
from chunklet.utils.error_utils import pretty_errors
from chunklet.models import CustomProcessorConfig 
from chunklet.exceptions import (
    InvalidInputError,
    UnsupportedFileTypeError,
    FileProcessingError,
)
from pydantic import ValidationError

class DocumentChunker:
    """
    A comprehensive document chunker that handles various file formats.

    This class provides a high-level interface to chunk text from different
    document types. It automatically detects the file format and uses the
    appropriate method to extract content before passing it to an underlying
    `PlainTextChunker` instance.

    Key Features:
    - Multi-Format Support: Chunks text from PDF, DOCX, TXT, MD, and RST files.
    - Tabular Data Handling: Provides a separate method (`chunk_table`) for
      intelligently chunking CSV and Excel files.
    - Metadata Enrichment: Automatically adds source file path and other
      document-level metadata (e.g., PDF page numbers) to each chunk.
    - Bulk Processing: Efficiently chunks multiple documents in a single call.
    - Pluggable Document processors: Integrate custom processors allowing definition
    of specific logic for extracting text from various file types.
    """
    # Supported file extensions for general-purpose text documents
    GENERAL_TEXT_EXTENSIONS = {".pdf", ".docx", ".txt", ".md", ".rst", ".rtf"}
    # Supported file extensions for tabular data (spreadsheets and CSVs)
    TABULAR_FILE_EXTENSIONS = {".xls", ".xlsx", ".csv"}

    def __init__(
        self,
        plain_text_chunker: PlainTextChunker,
        custom_processors: Optional[List[CustomProcessorConfig]] = None,
    ):
        """
        Initializes the DocumentChunker.

        Args:
            plain_text_chunker (PlainTextChunker): An instance of `PlainTextChunker` to be
                used for chunking the text extracted from documents.
            custom_processors (Optional[List[Dict]]): A list of custom document processors.
                Each processor should be a dictionary with
                'name' (str), 'File extensions' (Union[str, Iterable[str]])
                (e.g., '.json', ['.xml', '.json'])."),
                and a 'callback' (Callable[[str], List[str]])
                (Where the input is a path string) keys. 
        """
        if not isinstance(plain_text_chunker, PlainTextChunker):
            raise TypeError("plain_text_chunker must be an instance of PlainTextChunker")
        self.plain_text_chunker = plain_text_chunker

        if not custom_processors:
            custom_processors = {}
        try:
            self.custom_processors = [
                CustomProcessorConfig.model_validate(proc)
                for proc in custom_processors
            ]
        except ValidationError as e:
            pretty_err = pretty_errors(e)
            raise InvalidInputError(pretty_err)

        self._custom_processor_extensions = set()
        if self.custom_processors:
            for processor in self.custom_processors:
                if isinstance(processor.file_extensions, str):
                    self._custom_processor_extensions.add(processor.file_extensions)
                else:
                    self._custom_processor_extensions.update(processor.file_extensions)

        self.pdf_processor = PDFProcessor()
        self.docx_processor = DOCXProcessor()

    def validate_path(
        self,
        path: Union[str, Path],
        mode: str = "general"
    ) -> Tuple[Path, str]:
        """
        Validates the file path and its extension against supported types.

        Args:
            path (Union[str, Path]): The path to the document file.
            mode (str): The validation mode, either "general" for text documents or
                "table" for tabular data. Defaults to "general".

        Returns:
            Tuple[Path, str]: A tuple containing the resolved `Path` object and the
            file's lowercase extension.
        Raises:
            InvalidInputError: If the path is invalid or cannot be resolved.
            FileNotFoundError: If provided file path not found.
            ValueError: If the file extension is not supported for the given mode.
        """
        try:
            path = Path(str(path))
        except Exception as e:
            raise InvalidInputError from e

        extension = path.suffix.lower()
        if not extension:
            raise InvalidInputError(f"Invalid path: '{path}'. Path must have a recognizable extension.")

        if not path.is_file():
            raise FileNotFoundError(f"File not found: {path}")

        if mode == "general":
            supported_extensions = self.GENERAL_TEXT_EXTENSIONS | self._custom_processor_extensions
            if extension not in supported_extensions:
                raise UnsupportedFileTypeError(f"File type '{extension}' is not supported for general document chunking.")
        elif mode == "table":
            if extension not in self.TABULAR_FILE_EXTENSIONS:
                raise UnsupportedFileTypeError(f"File type '{extension}' is not supported for tabular data chunking.")

        return path, extension

    def _read_file(self, path: Union[str, Path], ext: str) -> str:
        """
        Reads the content of a text-based file and performs necessary conversions.

        This helper method handles reading from .txt, .md, and .rst files.
        It specifically converts .rst content to Markdown.

        Args:
            path (Union[str, Path]): The path to the file.
            ext (str): The file extension (e.g., ".txt", ".rst").

        Returns:
            str: The content of the file as a string.
        """
        if self.plain_text_chunker.verbose:
            logger.debug(f"Reading file: {path} with extension: {ext}")

        try:
            with open(path, "r", encoding="utf-8") as f:
                text_content = f.read()
        except (IOError, UnicodeDecodeError) as e:
            raise FileProcessingError(f"Error reading file {path}: {e}") from e
            
        if ext == ".rst":
            text_content = rst_to_markdown(text_content)
        elif ext == ".rtf":
            if rtf_to_text is None:
                raise FileProcessingError("The 'striprtf' library is not installed. Please install it with 'pip install striprtf' or install the document processing extras with 'pip install chunklet-py[document]'")
            text_content = rtf_to_text(text_content)

        return text_content

    def _use_custom_processor(self, path: Union[str, Path], ext: str) -> str:
        """
        Processes a file using a custom processor registered for the given extension.

        Args:
            path (Union[str, Path]): The path to the file.
            ext (str): The file extension (e.g., ".json").

        Returns:
            str: The extracted text content.

        Raises:
            InvalidInputError: If a custom processor fails during execution or if no
                               matching custom processor is found for the given extension.
        """
        for processor in self.custom_processors:
            supported_extensions = [processor.file_extensions] if isinstance(processor.file_extensions, str) else processor.file_extensions
            if ext in supported_extensions:
                if self.plain_text_chunker.verbose:
                    logger.debug(f"Using custom processor '{processor.name}' for file type: {ext}")
                try:
                    return processor.callback(str(path))
                except Exception as e:
                    raise FileProcessingError(f"Custom processor '{processor.name}' failed for file '{path}': {e}") from e

    def _create_chunk_boxes(
        self,
        chunks_str: List[str],
        base_metadata: Dict[str, Any]
    ) -> List[Box]:
        """
        Helper to create a list of Box objects for chunks with embedded metadata and auto-assigned chunk numbers.
        """
        final_chunks = []
        for i, chunk_str in enumerate(chunks_str):
            chunk_box = Box({
                "content": chunk_str,
                "metadata": {
                    "chunk_num": i + 1,
                }
            })
            chunk_box.metadata.update(base_metadata)
            final_chunks.append(chunk_box)
        return final_chunks

    def chunk(
        self,
        path: Union[str, Path],
        *,
        lang: str = "auto",
        mode: str = "sentence",
        max_tokens: int = 256,
        max_sentences: int = 12,
        overlap_percent: Union[int, float] = 20,
        offset: int = 0,
        token_counter: Optional[Callable[[str], int]] = None,
        n_jobs: Optional[int] = None, # only used for pdf
    ) -> List[Box]:
        """
        Chunks a single document from a given path.

        This method automatically detects the file type and uses the appropriate
        processor to extract text before chunking. It then adds document-level
        metadata to each resulting chunk.

        Args:
            path (Union[str, Path]): The path to the document file.
            lang (str): The language of the text. Defaults to "auto".
            mode (str): The chunking mode ('sentence', 'token', etc.).
            max_tokens (int): The maximum number of tokens per chunk.
            max_sentences (int): The maximum number of sentences per chunk.
            overlap_percent (Union[int, float]): The percentage of overlap.
            offset (int): The starting sentence offset for chunking.
            token_counter (Optional[Callable[[str], int]]): A custom token counter.
            n_jobs (Optional[int]): The number of parallel jobs (only used for PDFs).

        Returns:
            List[Box]: A list of `Box` objects, each representing a chunk with its
            content and metadata.

        Raises:
            InvalidInputError: If the path is invalid or cannot be resolved.
            FileNotFoundError: If provided file path not found.
            UnsupportedFileTypeError: If the file extension is not supported for the given mode.
            FileProcessingError: If an error occurs during file reading or processing.
            MissingTokenCounterError: If `mode` is "token" or "hybrid" but no `token_counter` is provided.
        """
        # Capture all parameters
        params = {k: v for k, v in locals().items() if k not in ['self', 'path', 'n_jobs']}

        validated_path, ext = self.validate_path(path, mode="general")
        
        if self.plain_text_chunker.verbose:
            logger.debug(f"Validated path: {validated_path}, detected extension: {ext}")

        text_content = "" # This will hold the text for chunking
        document_metadata = {"source":str(validated_path)}

        # Prioritize custom processors
        if ext in self._custom_processor_extensions:
            text_content = self._use_custom_processor(validated_path, ext)
            # Process as a single block of text
            chunks_str = self.plain_text_chunker.chunk(
                text=text_content,
                **params
            )

            final_chunks = self._create_chunk_boxes(chunks_str, document_metadata)

            if self.plain_text_chunker.verbose:
                logger.info(f"Generated {len(final_chunks)} chunks for {validated_path} using custom processor")

            return final_chunks

        if ext in {".txt", ".md", ".rst", ".docx", ".rtf"}: 
            if self.plain_text_chunker.verbose:
                logger.debug(f"Processing text-based file: {ext}")
            if ext == ".docx":
                text_content = self.docx_processor.extract_text(validated_path)
            else:
                text_content = self._read_file(validated_path, ext)
            
            # Process as a single block of text
            chunks_str = self.plain_text_chunker.chunk(
                text=text_content,
                **params
            )
            
            final_chunks = self._create_chunk_boxes(chunks_str, document_metadata)
            
            if self.plain_text_chunker.verbose:
                logger.info(f"Generated {len(final_chunks)} chunks for {validated_path}")

            return final_chunks

        elif ext == ".pdf":
            if self.plain_text_chunker.verbose:
                logger.debug("Processing PDF file with PDFProcessor")
            pdf_data = self.pdf_processor.extract_text(str(validated_path))
            
            # Prepare texts for plain_text_chunker.batch_chunk
            pages_text = pdf_data.pop("pages", [])
            document_metadata.update(pdf_data)
            
            # Call plain_text_chunker.batch_chunk
            params["show_progress"] = False # Dont show progress 
            all_pages_chunks = self.plain_text_chunker.batch_chunk(
                texts=pages_text,
                n_jobs=n_jobs,
                **params
            )

            final_chunks = []
            for page_num, chunks_str_list in enumerate(all_pages_chunks, start=1):
                page_metadata = document_metadata.copy()
                page_metadata["page_num"] = page_num
                final_chunks.extend(self._create_chunk_boxes(chunks_str_list, page_metadata))
            
            if self.plain_text_chunker.verbose:
                logger.info(f"Generated {len(final_chunks)} chunks for {validated_path} (from PDF)")

            return final_chunks
    
    def bulk_chunk(
        self,
        paths: List[Union[str, Path]],
        *,
        lang: str = "auto",
        mode: str = "sentence",
        max_tokens: int = 256,
        max_sentences: int = 12,
        overlap_percent: Union[int, float] = 20,
        offset: int = 0,
        token_counter: Optional[Callable[[str], int]] = None,
        n_jobs: Optional[int] = None,
        show_progress: bool = True, 
    ) -> List[List[Box]]:
        """
        Chunks multiple documents from a list of file paths in a single batch.

        This method extracts text from all documents, processes them together using
        the underlying `PlainTextChunker`'s `batch_chunk` method, and then groups
        the resulting chunks by their original document.

        Args:
            paths (List[Union[str, Path]]): A list of paths to the document files.
            lang (str): The language of the text.
            mode (str): The chunking mode.
            max_tokens (int): Maximum tokens per chunk.
            max_sentences (int): Maximum sentences per chunk.
            overlap_percent (Union[int, float]): Overlap percentage between chunks.
            offset (int): Token offset for subsequent chunks.
            token_counter (Optional[Callable[[str], int]]): Custom token counter.
            n_jobs (Optional[int]): Number of parallel jobs to use within each call
                to `chunk()` (e.g., for PDFs).
            show_progress (bool): Flag to show to enable or disable the loading bar.

        Returns:
            List[List[Box]]: A list of lists of `Box` objects, where each inner list contains
            chunks from a single processed document, in the same order as the input paths.

        Raises:
            InvalidInputError: If the path is invalid or cannot be resolved.
            FileNotFoundError: If provided file path not found.
            UnsupportedFileTypeError: If the file extension is not supported for the given mode.
            FileProcessingError: If an error occurs during file reading or processing.
            MissingTokenCounterError: If `mode` is "token" or "hybrid" but no `token_counter` is provided.
        """
        if self.plain_text_chunker.verbose:
            logger.info(f"Starting bulk chunking for {len(paths)} documents.")

        chunk_params = {k: v for k, v in locals().items() if k not in ['self', 'paths']}

        all_texts = []
        all_document_metadata = []
        doc_indices = []
        
        valid_paths_with_ext = []
        for path in paths:
            try:
                validated_path, ext = self.validate_path(path, mode="general")
                valid_paths_with_ext.append((validated_path, ext))
            except (FileNotFoundError, UnsupportedFileTypeError) as e:
                if self.plain_text_chunker.verbose:
                    logger.warning(f"Skipping file {path}: {e}")

        if not valid_paths_with_ext:
            return []

        for doc_index, (validated_path, ext) in enumerate(valid_paths_with_ext):
            document_metadata = {"source": str(validated_path)}

            if self.plain_text_chunker.verbose:
                logger.debug(f"Processing file: {validated_path}, detected extension: {ext}")

            # Prioritize custom processors
            if ext in self._custom_processor_extensions:
                text_content = self._use_custom_processor(validated_path, ext)
                all_texts.append(text_content)
                all_document_metadata.append(document_metadata)
                doc_indices.append(doc_index)
                
            elif ext == ".pdf":
                pdf_data = self.pdf_processor.extract_text(str(validated_path))
                pages_text = pdf_data.pop("pages", [])
                document_metadata.update(pdf_data)
                all_texts.extend(pages_text)
                
                for i in range(len(pages_text)):
                    page_metadata = document_metadata.copy()
                    page_metadata["page_num"] = i + 1
                    all_document_metadata.append(page_metadata)
                    doc_indices.append(doc_index)
                    
            elif ext == ".docx":
                text_content = self.docx_processor.extract_text(validated_path)
                all_texts.append(text_content)
                all_document_metadata.append(document_metadata)
                doc_indices.append(doc_index)
                
            elif ext in self.GENERAL_TEXT_EXTENSIONS - {".pdf", ".docx"}:
                text_content = self._read_file(validated_path, ext)
                all_texts.append(text_content)
                all_document_metadata.append(document_metadata)
                doc_indices.append(doc_index)

        if not all_texts:
            if self.plain_text_chunker.verbose:
                logger.info("No valid texts found for batch chunking. Returning empty list.")
            return []

        all_chunks_from_plain_text_chunker = self.plain_text_chunker.batch_chunk(
            texts=all_texts, **chunk_params
        )

        final_results_per_document = [[] for _ in valid_paths_with_ext]
        for i, doc_chunks_str_list in enumerate(all_chunks_from_plain_text_chunker):
            doc_index = doc_indices[i]
            curr_doc_metadata = all_document_metadata[i]

            # Use the helper to create chunk boxes for this document
            chunks_for_this_doc = self._create_chunk_boxes(doc_chunks_str_list, curr_doc_metadata)
            final_results_per_document[doc_index].extend(chunks_for_this_doc)

        if self.plain_text_chunker.verbose:
            logger.info(
                f"Finished bulk chunking. Generated {len(final_results_per_document)} document chunk lists from {len(valid_paths_with_ext)} documents."
            )
        return final_results_per_document

    def chunk_table(
        self,
        path: Union[str, Path],
        *,
        max_lines: int = 100,
    ) -> List[Box]:
        """
        Chunks a tabular data file (Excel or CSV) into smaller tables.

        Each chunk is a list of lists, with the header row repeated at the top
        of each chunk. This method is designed for structured tabular data.

        Args:
            path (Union[str, Path]): The path to the Excel or CSV file.
            max_lines (int): The maximum number of lines (header + data rows)
                per chunk. Must be greater than 1.

        Returns:
            List[Box]: A list of `Box` objects, where each chunk's content is a
            smaller table (list of lists).

        Raises:
            FileNotFoundError: If the specified path does not exist.
            InvalidInputError: If `max_lines` is not greater than 1.
            FileProcessingError: If there is an error processing the file.
        """
        try:
            import openpyxl
        except ImportError:
            raise FileProcessingError(
                "The 'openpyxl' library is not installed. "
                "Please install it with 'pip install openpyxl' or install the document processing extras "
                "with 'pip install 'chunklet-py[document]''"
            )
        if self.plain_text_chunker.verbose:
            logger.info(f"Starting tabular chunking for {path}")

        path, file_extension = self.validate_path(path, mode="table")
        
        document_metadata: Dict[str, Any] = {"source":str(path)}
        rows: List[List[Any]] = []

        try:
            if file_extension in {".xls", ".xlsx"}:
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active 
                rows = [list(row) for row in sheet.iter_rows(values_only=True) if any(row)]
                
                if workbook.properties.creator:
                    document_metadata["author"] = workbook.properties.creator
                if workbook.properties.title:
                    # Add active sheet name
                    document_metadata["title"] = workbook.properties.title
                document_metadata["sheet_name"] = sheet.title
                document_metadata["sheet_count"] = len(workbook.sheetnames)
            elif file_extension == '.csv':
                with open(path, "r", encoding="utf-8", newline='') as f:
                    reader = csv.reader(f)
                    rows = [list(row) for row in reader]
            
            if not rows:
                if self.plain_text_chunker.verbose:
                    logger.info(f"No data found in {path}. Returning empty list.")
                return []

            header, *data_rows = rows
            
        except Exception as e:
            raise FileProcessingError(f"Error processing file {path}: {e}") from e
            
        chunks = []
        # chunk_size is the number of data rows per chunk
        chunk_size = max_lines - 1 
        if chunk_size <= 0:
            raise InvalidInputError("max_lines must be greater than 1 to include the header and at least one data row.")

        for i in range(0, len(data_rows), chunk_size):
            chunk_box = Box()
            chunk_box.content = [header] + data_rows[i:i + chunk_size]
            chunk_box.metadata = document_metadata.copy()
            chunk_box.metadata.start_row = i + 2 # 1-based index + header
            chunk_box.max_lines = max_lines 
            chunks.append(chunk_box)
            
        if self.plain_text_chunker.verbose:
            logger.info(f"Generated {len(chunks)} chunks for tabular file {path}")

        return chunks
        