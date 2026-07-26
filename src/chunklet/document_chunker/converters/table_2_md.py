import csv
from pathlib import Path

# openpyxyl is lazy imported

from chunklet.document_chunker.md_table import build_md_table


def table_to_md(file_path: str | Path) -> str:
    """
    Convert a CSV or XLSX file into a Markdown-formatted table string.

    Args:
        file_path: Path to the input file (.csv or .xlsx).

    Returns:
        Markdown table representation of the file contents.
    """
    file_path = Path(file_path)
    ext = file_path.suffix.lower()

    if ext == ".csv":
        with open(file_path, newline="", encoding="utf-8") as f:
            data = list(csv.reader(f))
    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover
            raise ImportError(
                "The 'openpyxl' library is not installed. "
                "Please install it with 'pip install openpyxl>=3.1.2' "
                "or install the document processing extras with "
                "'pip install chunklet-py[structured-document]'"
            ) from e
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        data = list(sheet.iter_rows(values_only=True))
        wb.close()
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return build_md_table(data)


# --- Example usage ---
if __name__ == "__main__":  # pragma: no cover
    sample_file = "samples/example.xlsx"
    md_table = table_to_md(sample_file)
    print(f"\nMarkdown output for {sample_file}:\n")
    print(md_table)
